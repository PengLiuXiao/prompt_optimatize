"""Dataset loading and input validation. Pure format checks, no inference."""
from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any


INPUT_COLUMNS = {"input", "text", "query", "utterance", "样本", "输入", "文本", "用户输入"}
GOLD_COLUMNS = {"gold", "label", "labels", "answer", "expected", "正确答案", "标注", "标签"}


def load_dataset(
    path: str | Path,
    input_column: str | None = None,
    gold_column: str | None = None,
    strip_gold: bool = False,
) -> list[dict[str, Any]]:
    """Load and normalize a dataset to rows of {id, input, gold}.

    If strip_gold=True, output only {id, input} — used to prepare inputs
    for run-inference without exposing gold answers.
    """
    dataset_path = Path(path)
    suffix = dataset_path.suffix.lower()
    if suffix == ".jsonl":
        rows = _load_jsonl(dataset_path)
    elif suffix == ".csv":
        rows = _normalize_tabular(_load_csv(dataset_path), input_column, gold_column)
    elif suffix in {".xlsx", ".xlsm"}:
        rows = _normalize_tabular(_load_excel(dataset_path), input_column, gold_column)
    else:
        raise ValueError(f"unsupported dataset file type: {suffix}")

    if strip_gold:
        return [{"id": row["id"], "input": row["input"]} for row in rows]
    return rows


def validate_inputs(
    path: str | Path,
    input_column: str | None = None,
    gold_column: str | None = None,
    sample_size: int = 5,
) -> dict[str, Any]:
    """Validate dataset format and gold-answer presence.

    Returns {valid, row_count, sample_rows, error}. Does NOT infer task_type;
    that is the Agent's job in the eval-plan step.
    """
    try:
        rows = load_dataset(path, input_column=input_column, gold_column=gold_column)
    except (FileNotFoundError, ValueError) as error:
        return {"valid": False, "row_count": 0, "sample_rows": [], "error": str(error)}

    if not rows:
        return {"valid": False, "row_count": 0, "sample_rows": [], "error": "dataset is empty"}

    missing = [
        str(row.get("id", index))
        for index, row in enumerate(rows, start=1)
        if "gold" not in row or _is_empty_gold(row.get("gold"))
    ]
    if missing:
        return {
            "valid": False,
            "row_count": len(rows),
            "sample_rows": [],
            "error": f"dataset must include a human gold answer for every row; missing gold answer on row(s): {', '.join(missing)}",
        }

    return {
        "valid": True,
        "row_count": len(rows),
        "sample_rows": rows[:sample_size],
        "error": None,
    }


def _is_empty_gold(value: Any) -> bool:
    return value is None or value == "" or value == [] or value == {}


def _load_jsonl(path: Path) -> list[dict[str, Any]]:
    rows = []
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            if line.strip():
                row = json.loads(line)
                if "gold" in row:
                    row["gold"] = _parse_gold(row["gold"])
                rows.append(row)
    return rows


def _load_csv(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def _load_excel(path: Path) -> list[dict[str, Any]]:
    import openpyxl

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    records = []
    for values in rows[1:]:
        record = {}
        for header, value in zip(headers, values):
            record[header] = "" if value is None else value
        if any(str(value).strip() for value in record.values()):
            records.append(record)
    return records


def _normalize_tabular(
    records: list[dict[str, Any]],
    input_column: str | None,
    gold_column: str | None,
) -> list[dict[str, Any]]:
    if not records:
        return []
    headers = list(records[0].keys())
    input_col = input_column or _detect_column(headers, INPUT_COLUMNS, "input")
    gold_col = gold_column or _detect_column(headers, GOLD_COLUMNS, "gold")
    normalized = []
    for index, record in enumerate(records, start=1):
        gold = _parse_gold(record.get(gold_col))
        normalized.append(
            {
                "id": str(record.get("id") or record.get("ID") or index),
                "input": str(record.get(input_col) or ""),
                "gold": gold,
            }
        )
    if any(isinstance(row["gold"], list) for row in normalized):
        normalized = [
            dict(row, gold=row["gold"] if isinstance(row["gold"], list) else [row["gold"]])
            for row in normalized
        ]
    return normalized


def _detect_column(headers: list[str], candidates: set[str], purpose: str) -> str:
    matches = [header for header in headers if header.strip() in candidates]
    if not matches:
        raise ValueError(f"cannot detect {purpose} column; pass --{purpose}-column")
    if len(matches) > 1:
        raise ValueError(f"ambiguous {purpose} columns: {', '.join(matches)}")
    return matches[0]


def _parse_gold(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (list, dict, int, float)):
        return value
    text = str(value).strip()
    if not text:
        return ""
    if (text.startswith("{") and text.endswith("}")) or (text.startswith("[") and text.endswith("]")):
        try:
            return json.loads(text)
        except json.JSONDecodeError:
            pass
    try:
        if "." in text:
            return float(text)
        return int(text)
    except ValueError:
        pass
    separators = [",", "，", ";", "；"]
    if any(separator in text for separator in separators):
        parts = [text]
        for separator in separators:
            next_parts = []
            for part in parts:
                next_parts.extend(part.split(separator))
            parts = next_parts
        return [part.strip() for part in parts if part.strip()]
    return text
